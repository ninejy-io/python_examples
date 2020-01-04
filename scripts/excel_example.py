import datetime
import openpyxl


# wb = openpyxl.Workbook()
# ws = wb.active

# title = ['Name', 'Age', 'Like', 'Job', 'Now']
# row1 = ['Robert', 18, 'reading', 'student', datetime.datetime.now()]

# ws.append(title)
# ws.append(row1)

# wb.save('sample.xlsx')


wb = openpyxl.load_workbook('sample.xlsx')
ws = wb.active

# rows = ws.rows
# for y in rows:
#     for x in range(len(y)):
#         print(y[x].value)

# columns = ws.columns
# for x in columns:
#     for y in range(len(x)):
#         print(x[y].value)

# iter_rows = ws.iter_rows()
# for i in iter_rows:
#     print(i)

# iter_cols = ws.iter_cols()
# for j in iter_cols:
#     print(j)

values = ws.values
for line in values:
    print(line)
